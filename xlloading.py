import pandas as pd
from openpyxl import load_workbook

# Path to your existing Excel file
file_path = 'sample.xlsx'

# Create new data
new_data = pd.DataFrame({
    'Name': ['Virat', 'Rohit'],
    'Score': [92, 85]
})

# Load the workbook to get current row
book = load_workbook(file_path)
sheet = book['sheet1']  # Match exact sheet name (case-sensitive)
startrow = sheet.max_row

# Write to Excel file in append mode (no need to assign writer.sheets!)
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    new_data.to_excel(writer, sheet_name='sheet1', startrow=startrow, index=False, header=True)
